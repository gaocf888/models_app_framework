from __future__ import annotations

from html.parser import HTMLParser
from pathlib import Path


class _TextExtractor(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self._parts: list[str] = []

    def handle_data(self, data: str) -> None:
        if data and data.strip():
            self._parts.append(data.strip())

    def text(self) -> str:
        return "\n".join(self._parts)


class DocumentParser:
    def parse(self, content: str, source_type: str) -> str:
        st = (source_type or "text").lower()
        if st in {"text", "txt"}:
            return content
        if st in {"md", "markdown"}:
            return self._parse_markdown(content)
        if st == "html":
            return self._parse_html(content)
        if st == "pdf":
            return self._parse_pdf(content)
        if st in {"docx", "doc"}:
            return self._parse_docx(content)
        if st in {"xlsx", "xlsm"}:
            return self._parse_xlsx(content)
        return content

    @staticmethod
    def resolve_local_path(content: str) -> Path | None:
        raw = (content or "").strip()
        if not raw:
            return None
        # 支持 file:// 前缀，或直接传本地绝对路径
        if raw.lower().startswith("file://"):
            raw = raw[7:]
        p = Path(raw)
        if p.exists() and p.is_file():
            return p
        return None

    @staticmethod
    def _parse_markdown(content: str) -> str:
        # 保留标题与正文，去除 code fence 标记符本身
        lines = []
        in_fence = False
        for line in content.splitlines():
            if line.strip().startswith("```"):
                in_fence = not in_fence
                continue
            lines.append(line)
        return "\n".join(lines)

    @staticmethod
    def _parse_html(content: str) -> str:
        parser = _TextExtractor()
        parser.feed(content)
        return parser.text()

    @staticmethod
    def _parse_pdf(content: str) -> str:
        p = DocumentParser.resolve_local_path(content)
        if p is None:
            # 兼容旧行为：上游已传提取后的纯文本
            return content
        try:
            from pypdf import PdfReader  # type: ignore[import-untyped]
        except Exception as e:  # noqa: BLE001
            raise ImportError("Parsing PDF path requires pypdf. Install from requirements.") from e
        reader = PdfReader(str(p))
        parts: list[str] = []
        for page in reader.pages:
            txt = page.extract_text() or ""
            if txt.strip():
                parts.append(txt.strip())
        return "\n\n".join(parts).strip()

    @staticmethod
    def _serialize_docx_table(table: object) -> str:
        """将 Word 表格转为带标记的纯文本，便于 RAG 分块与关键词命中。"""
        rows = list(table.rows)
        if not rows:
            return ""
        col_counts = [len(r.cells) for r in rows]
        ncols = max(col_counts) if col_counts else 0
        header = f"[DOCX_TABLE rows={len(rows)} cols={ncols}]"
        lines: list[str] = [header]
        for row in rows:
            cells: list[str] = []
            for cell in row.cells:
                txt = (cell.text or "").replace("\r", " ").replace("\n", " ").strip()
                txt = txt.replace("|", "｜")
                cells.append(txt)
            lines.append(" | ".join(cells))
        return "\n".join(lines).strip()

    @staticmethod
    def _parse_docx(content: str) -> str:
        """
        解析 DOCX：按文档顺序输出段落与表格（旧实现仅段落，表格会丢失）。

        说明：.doc 老格式需用户自行转换为 .docx；python-docx 不支持二进制 .doc。
        """
        p = DocumentParser.resolve_local_path(content)
        if p is None:
            return content
        try:
            import docx  # type: ignore[import-untyped]
        except Exception as e:  # noqa: BLE001
            raise ImportError("Parsing DOCX path requires python-docx. Install from requirements.") from e
        from docx.oxml.ns import qn
        from docx.table import Table
        from docx.text.paragraph import Paragraph

        d = docx.Document(str(p))
        segments: list[str] = []
        for child in d.element.body:
            if child.tag == qn("w:p"):
                para = Paragraph(child, d)
                t = para.text.strip()
                if t:
                    segments.append(t)
            elif child.tag == qn("w:tbl"):
                tbl = Table(child, d)
                tbl_text = DocumentParser._serialize_docx_table(tbl)
                if tbl_text:
                    segments.append(tbl_text)
        return "\n\n".join(segments).strip()

    @staticmethod
    def _parse_xlsx(content: str) -> str:
        """
        解析 Excel（.xlsx / .xlsm）：逐工作表展开为纯文本行，供向量化与检索。

        说明：传统二进制 .xls 不在此支持，需转换为 xlsx 或另接 xlrd。
        """
        p = DocumentParser.resolve_local_path(content)
        if p is None:
            return content
        try:
            from openpyxl import load_workbook  # type: ignore[import-untyped]
        except Exception as e:  # noqa: BLE001
            raise ImportError("Parsing XLSX path requires openpyxl. Install from requirements.") from e

        parts: list[str] = []
        wb = load_workbook(filename=str(p), read_only=True, data_only=True)
        try:
            for sheet in wb.worksheets:
                parts.append(f"[XLSX_SHEET name={sheet.title}]")
                for row in sheet.iter_rows(values_only=True):
                    cells = [
                        "" if c is None else str(c).replace("\r", " ").replace("\n", " ").strip()
                        for c in row
                    ]
                    if not any(cells):
                        continue
                    parts.append(" | ".join(c.replace("|", "｜") for c in cells))
                parts.append("")
        finally:
            wb.close()
        return "\n".join(parts).strip()
